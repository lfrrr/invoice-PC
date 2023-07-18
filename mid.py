import re
import openpyxl

from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine
from openpyxl import load_workbook

from pathlib import Path
import pandas as pd
import sys
import importlib
import os
import random
import shutil
from PyPDF2 import PdfReader,PdfMerger
import time

prefix = "_未报销"
prefix1 = "_已报销"
suffix = ".pdf"
kinds = ['已报销', '未报销']
types = ['交通费', '差旅费', '耗材费', '印刷费', '培训费', '其他']


# strtime = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')


# 解析PDF文件
def parse_pdf(path, output_path):
    with open(path, 'rb') as fp:
        parser = PDFParser(fp)
        doc = PDFDocument()
        parser.set_document(doc)
        doc.set_parser(parser)
        doc.initialize('')
        rsrcmgr = PDFResourceManager()
        laparams = LAParams(all_texts=True, boxes_flow=2.0, heuristic_word_margin=True)
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        extracted_text = ''
        for page in doc.get_pages():
            interpreter.process_page(page)
            layout = device.get_result()
            for lt_obj in layout:
                if isinstance(lt_obj, LTTextBox) or isinstance(lt_obj, LTTextLine):
                    extracted_text += lt_obj.get_text()
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(extracted_text)


# 标题关键字
def inner_key_word():
    return ["发票代码", "发票号码", "开票日期", "货物或应税劳务、服务名称", "金额（含税）", "票据类别"]


# 转换汉字和数字
CN_NUM = {
    '〇': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '零': 0,
    '壹': 1, '贰': 2, '叁': 3, '肆': 4, '伍': 5, '陆': 6, '柒': 7, '捌': 8, '玖': 9, '貮': 2, '两': 2,
}

CN_UNIT = {
    '分': 0.01,
    '角': 0.1,
    '元': 1.00,
    '圆': 1.00,
    '整': 1.00,
    '十': 10,
    '拾': 10,
    '百': 100,
    '佰': 100,
    '千': 1000,
    '仟': 1000,
    '万': 10000,
    '萬': 10000,
    '亿': 100000000,
    '億': 100000000,
}


def chinese_to_arabic(cn):
    unit = 0  # current
    ldig = []  # digest
    for cndig in reversed(cn):
        if cndig in CN_UNIT:
            unit = CN_UNIT.get(cndig)
            if unit == 10000 or unit == 100000000:
                ldig.append(unit)
                unit = 1
        else:
            dig = CN_NUM.get(cndig)
            if unit:
                dig *= unit
                unit = 0
            ldig.append(dig)
    if unit == 10:
        ldig.append(10)
    val, tmp = 0, 0
    for x in reversed(ldig):
        if x == 10000 or x == 100000000:
            val += tmp * x
            tmp = 0
        else:
            tmp += x
    val += tmp
    return val


# 正则化
# 将text内容按行写入list
# 遍历list，从中选择必要内容：
# 发票代码： 12位数字，且上下文没有“机器编号”字样
# 发票号码： 8位数字
# 开票日期： 四位数字 + 一位 空格/“年” + 两位数字 + 一位 空格/“月” + 两位数字（字符型）
# 货物或应税劳务、服务名称： * + 汉字 + * + 汉字
# 金额（含税）： 含有汉字大写数字
# 类别：等到具体时再写入
def split_item(text):
    ITEMS = [None] * 6  # 设计一个list类型
    text_line = text.split("\n")  # 将text内容按行写入一个表格
    licnt = len(text_line) - 1  # 行数减一
    for index, item in enumerate(text_line):
        if re.match(r'\d\d\d\d\d\d\d\d\d\d\d\d$', item):  # 是12位数字 发票代码
            if index == 0 and re.sub(r"\s+", "", text_line[index + 1]) == "机器编号:":
                continue
            elif licnt > index >= 1 and (re.sub(r"\s+", "", text_line[index + 1]) == "机器编号:" or re.sub(r"\s+", "",
                                                                                                       text_line[
                                                                                                           index - 1]) == "机器编号:"):
                continue
            elif index == licnt and re.sub(r"\s+", "", text_line[index - 1]) == "机器编号:":
                continue
            else:
                ITEMS[0] = item
        if re.match(r'\d\d\d\d\d\d\d\d$', item):  # 如果是8位数字 发票号码
            ITEMS[1] = item
        if re.match(r'20\d\d\D+\d\d\D+\d\d', item):  # 是年月日 日期
            ITEMS[2] = item
        if re.match(r'\*[\u4e00-\u9fa5]+\*', item):  # 如果符合 * + 汉字 + * + 汉字 格式 部分货物劳务名称
            if ITEMS[3] is None:  # 如果为空
                ITEMS[3] = item
            else:  # 如果含有多件商品
                ITEMS[3] = ITEMS[3] + item
        if re.match(r'[\u58f9\u8d30\u53c1\u8086\u4f0d\u9646\u67d2\u634c\u7396]', item):  # 如果开头含有汉字大写数字中任意一个，金额 用大写字母转
            # 转换汉字和数字
            itint = chinese_to_arabic(item)
            ITEMS[4] = itint

    return ITEMS


'''
计划功能：
判断是否存在目标Excel文件，不存在则新建，存在则打开
读取当前文件行数，录入时插入最后一行
对于每一个key_word，匹配，再录入
'''


def write_to_excel(output_path, Items, key_words):  # output_path: 文件路径  Items: 内容块  key_words: 列标题
    if not os.path.exists(output_path):  # 如果不存在此文件
        wb = openpyxl.Workbook()  # 创建新文件
        sheet = wb.active  # 获取当前活跃页
        sheet.append(key_words)  # 将表头写入
        sheet.append(Items)  # 在末尾追加数据  此处 Items 应为 list 类型
        wb.save(output_path)  # 保存
    else:
        wb = load_workbook(output_path)  # 如果存在，打开该文件
        sheet = wb["Sheet"]  # 打开 Sheet
        sheet.append(Items)  # 在末尾追加数据  此处 Items 应为 list 类型
        wb.save(output_path)  # 保存


def seek_creat_files(id1, name):
    """根据输入的文件名称查找对应的文件夹有无文对应文件，无则输出创建文件"""
    flag = 0
    for root in os.listdir(id1):  # 只遍历第一层目录
        # , dirs, files  os.walk(id1)
        if name in root:
            # 当层文件内有该文件，跳出循环并返回文件地址
            id1 = id1 + '/' + name
            flag = 1
            break
    if flag == 0:  # 文件夹中不存在该文件夹，则创建
        id1 = id1 + '/' + name
        os.mkdir(id1)
    return id1


# 查找指定位置的文件，有则删除，无则返回交互语言
def removefile(fileaddress):
    if os.path.exists(fileaddress):  # 如果文件存在
        # 删除文件，可使用以下两种方法。
        os.remove(fileaddress)
    # os.unlink(path)
    else:
        #     return 0
        return 'no such file'  # 则返回文件不存在


# 实现excel文件信息去重
def throw_repeat(filepath, sheetname):
    importlib.reload(sys)
    # 读取Excel中Sheet1中的数据
    data = pd.DataFrame(pd.read_excel(filepath, sheetname, dtype=str))
    # 查看是否有重复行
    re_row = data.duplicated()
    # 查看去除重复行的数据
    no_re_row = data.drop_duplicates()
    # 查看基于['发票号码']列去除重复行的数据
    wp = data.drop_duplicates(subset=['发票号码'])
    # 将原来的excel删除，然后新建新的输入
    removefile(filepath)
    wp.to_excel(filepath)
    # 除掉重新写入的文件中的第一列，统一格式
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.delete_cols(1)  # 删除第 1 列数据
    wb.save(filepath)


# 实现excel内部数据按照金额排序
def excel_sort(filepath, sheetname):
    data_test = pd.read_excel(filepath)
    df = pd.DataFrame(data_test)
    # 以列“Fruit”的标签列来进行升序排列
    df_1 = df.sort_values('金额（含税）', ascending=True)
    writer = pd.ExcelWriter(filepath)
    df_1.to_excel(writer, sheet_name=sheetname, index=False)
    writer.save()


# 判断文件夹是否存在，文件路径不存在则创建文件夹
def check_and_creat_dir(file_url):
    '''
    :param file_url: 文件路径，包含文件名
    :return:
    '''
    file_gang_list = file_url.split('/')
    if len(file_gang_list) > 1:
        if not os.path.exists(file_url):
            os.makedirs(file_url)
        else:
            return None
        # 还可以直接创建空文件
    else:
        return None


# 对于未报销的发票文件夹，创建excel,去重，排序,同时实现发票分类存储
def UnuesdTicket_produceexcel_throw_sort(unuedfilepath, excelsiorpath, typenum):
    files = os.listdir(unuedfilepath)  # 得到文件夹下的所有文件名称
    s = []
    strtemp = 0
    newpath = unuedfilepath + "/未报销"
    # 创建5个类别的文件夹
    check_and_creat_dir(newpath + os.sep + "交通费")
    check_and_creat_dir(newpath + os.sep + "差旅费")
    check_and_creat_dir(newpath + os.sep + "耗材费")
    check_and_creat_dir(newpath + os.sep + "印刷费")
    check_and_creat_dir(newpath + os.sep + "培训费")
    check_and_creat_dir(newpath + os.sep + "其他")
    for file in files:  # 遍历unuedfilepath文件夹
        path = unuedfilepath + os.sep + file  # unuedfilepath下每个文件（夹）路径
        if not os.path.isdir(path):  # 判断是否是文件夹，不是文件夹才打开
            if os.path.splitext(file)[1] == ".pdf":  # 再判断是否为pdf文件，是再打开
                rpdf = unuedfilepath + "/" + file
                rtxt = unuedfilepath + "/temp.txt"  # 将转换所需的txt文件创建在pdf发票的目录下
                parse_pdf(rpdf, rtxt)
                text = open(rtxt, 'r', encoding='utf-8').read()

                key_words = inner_key_word()
                ITEMS = split_item(text)
                # for index, value in enumerate(ITEMS):
                #     print(value)
                value = str(ITEMS[0])
                # 货物或应税劳务、服务名称
                value1 = str(ITEMS[3])

                # todo:在这里验证是否已经报销过了
                # 移动未报销发票进入文件夹

                # os.sep为根据系统不同产生的不同的分隔符
                newfile = value + prefix + suffix
                new_name = unuedfilepath + os.sep + newfile
                repeat = newpath + os.sep + newfile
                removefile(repeat)  # 存在重复发票，则删除；不存在则返回 no such file

                old_name = unuedfilepath + os.sep + file
                os.rename(old_name, new_name)
                shutil.move(new_name, newpath)

                new_name = newpath + os.sep + newfile
                # 根据类别将发票移动到对应文件夹中
                # types = ['交通费', '差旅费', '耗材费', '印刷费', '培训费', '其他']
                if value1[0:5] == "*运输服务":
                    shutil.move(new_name, newpath + os.sep + types[0])
                    ITEMS[5] = types[0]
                    type = 0
                elif value1[0:5] == "*餐饮服务" or value1[0:5] == "*住宿服务":
                    shutil.move(new_name, newpath + os.sep + types[1])
                    ITEMS[5] = types[1]
                    type = 1
                elif value1[0:7] == "*移动通信设备" or value1[0:8] == "*计算机网络设备" or value1[0:8] == "*计算机配套产品" or value1[
                                                                                                           0:9] == "*复印胶版印制设备" or \
                        value1[0:3] == "*文具" or value1[0:5] == "*塑料制品" or value1[0:10] == "*仪器仪表办公用机械" or value1[
                                                                                                          0:4] == "*纸制品" or value1[
                                                                                                                            0:6] == "*眼镜类产品":
                    shutil.move(new_name, newpath + os.sep + types[2])
                    ITEMS[5] = types[2]
                    type = 2
                elif value1[0:4] == "*印刷品":
                    shutil.move(new_name, newpath + os.sep + types[3])
                    ITEMS[5] = types[3]
                    type = 3
                elif value1[0:5] == "*生活服务":
                    shutil.move(new_name, newpath + os.sep + types[4])
                    ITEMS[5] = types[4]
                    type = 4
                else:
                    shutil.move(new_name, newpath + os.sep + types[5])
                    ITEMS[5] = types[5]
                    type = 5
                # 判断此发票是否是用户需要报销的类别，若是则将其存入数据库
                for n in typenum:
                    if type == n:
                        write_to_excel(excelsiorpath, ITEMS, key_words)  # 可以输入不存在的文件，将会新建这个文件
                        throw_repeat(excelsiorpath, "Sheet")  # 去重
                        excel_sort(excelsiorpath, "Sheet")  # 排序

                removefile(rtxt)
        elif file == "未报销":
            for n in typenum:
                select_path = newpath + os.sep + types[n]  # 用户想要报销的类别对应的文件夹
                files = os.listdir(select_path)  # 得到文件夹下的所有文件名称
                for file in files:  # 遍历文件夹，解析发票
                    path = select_path + os.sep + file
                    if not os.path.isdir(file):  # 判断是否是文件夹，不是文件夹才打开
                        if os.path.splitext(file)[1] == ".pdf":  # 再判断是否为pdf文件，是再打开
                            rpdf = select_path + "/" + file
                            rtxt = select_path + "/temp.txt"  # 将转换所需的txt文件创建在pdf发票的目录下
                            parse_pdf(rpdf, rtxt)
                            text = open(rtxt, 'r', encoding='utf-8').read()
                            key_words = inner_key_word()
                            ITEMS = split_item(text)
                            ITEMS[5] = types[n]

                            write_to_excel(excelsiorpath, ITEMS, key_words)  # 可以输入不存在的文件，将会新建这个文件
                            throw_repeat(excelsiorpath, "Sheet")  # 去重
                            excel_sort(excelsiorpath, "Sheet")  # 排序
                            removefile(rtxt)


def Update_excel(WbxFilePath, excelsiorpath):
    removefile(excelsiorpath)
    files = os.listdir(WbxFilePath)  # 得到文件夹下的所有文件名称

    for file in files:  # 遍历文件夹
        if not os.path.isdir(file):  # 判断是否是文件夹，不是文件夹才打开
            if os.path.splitext(file)[1] == ".pdf":  # 再判断是否为pdf文件，是再打开
                rpdf = WbxFilePath + "/" + file
                rtxt = WbxFilePath + "/temp.txt"  # 将转换所需的txt文件创建在pdf发票的目录下
                parse_pdf(rpdf, rtxt)
                text = open(rtxt, 'r', encoding='utf-8').read()

                key_words = inner_key_word()
                ITEMS = split_item(text)

                write_to_excel(excelsiorpath, ITEMS, key_words)  # 可以输入不存在的文件，将会新建这个文件
                throw_repeat(excelsiorpath, "Sheet")  # 去重
                excel_sort(excelsiorpath, "Sheet")  # 排序

                removefile(rtxt)


# 利用组合的蒙特卡洛方法组合
def excel_MC_combination(UnuseFilepath,excelsiorpath, suma):
    '''读取表格中的发票形成列表'''
    stexcel = pd.read_excel(excelsiorpath)

    ex_value = stexcel.values
    len_r = len(ex_value)
    data = []
    for i in range(len_r):
        data.append(ex_value[i][4])  # 将所有金额形成列表

    '''利用组合的蒙特卡洛方法进行最优组合'''
    suma = float(suma)
    lendata = len(data)
    data1 = []
    res_data = []  # 存放最佳组合
    res_st = 0.0  # 组合加和最大值
    data1[:] = data[:]
    res_err = 10000.0
    flag = 0  # 用来判断是否有可以报销的发票
    for i in range(20000):  # 把0到19999依次赋给i，即重复20000次
        k = random.randrange(1, lendata + 1)  # 从给定的范围返回随机项（1，lendata+1)
        data2 = random.sample(data1, k)  # 从list中随机获取k个元素，作为一个片断返回赋给data2
        # print('i=',i,'k=',k)
        # print('data2=',data2)
        sumt = sum(data2)
        ds = suma - sumt  # ds为给定值和组合最大值的差
        if ds < 0:
            continue
        else:
            flag = 1
            if (ds < res_err):
                # print('')
                # print('ds=',ds)
                # print('i=',i,'data2=',data2)
                # print('k=',k)
                # print('sum=',sumt)
                res_err = ds
                res_data = data2
                res_st = sumt
                if (ds < 0.0000001): break
    if (flag == 0):
        list="根据给定金额，暂无可以报销的发票！"
        num=0
    else:
        list="在给定值下金额的组合为：" + str(sorted(res_data))
        list=list+"总金额为："+str(round(res_st, 2))+" 总金额与给定值相差："+str(round(res_err, 2))+"\n"
        list=list+"需要报销的发票最优组合如下："
        num=1
        selected_path = UnuseFilepath + "/selected"
        check_and_creat_dir(selected_path)
        # 清除此文件夹中已有的文件
        del_list = os.listdir(selected_path)
        for f in del_list:
            file_path = os.path.join(selected_path, f)
            os.remove(file_path)

        len2 = len(res_data)

        for i in range(len2):
            ind = data.index(res_data[i])  # 求data2的所有元素在data中的下标
            list=list+ str(ex_value[ind]) # 输出要报销发票的组合

            # os.sep为根据系统不同产生的不同的分隔符
            new_path = UnuseFilepath + "/已报销"
            new_name = UnuseFilepath + "/未报销" + os.sep + str(ex_value[ind][5]) + os.sep + "0" + str(
                ex_value[ind][0]) + prefix1 + suffix
            repeat = new_path + "/未报销" + os.sep + str(ex_value[ind][5]) + os.sep + "0" + str(
                ex_value[ind][0]) + prefix1 + suffix
            removefile(repeat)  # 存在重复发票，则删除；不存在则返回 no such file

            old_name = UnuseFilepath + "/未报销" + os.sep + str(ex_value[ind][5]) + os.sep + "0" + str(
                ex_value[ind][0]) + prefix + suffix
            os.rename(old_name, new_name)
            shutil.copy(new_name, selected_path)  # 将需要报销的发票放入"selected"文件夹中
            shutil.move(new_name, new_path)

            # 更新excel
            Update_excel(UnuseFilepath + "/未报销" + os.sep + str(ex_value[ind][5]), excelsiorpath)
        import glob
        file_names = os.listdir(selected_path)
        file_paths = glob.glob(os.path.join(selected_path, '*.pdf'))

        file_merger = PdfMerger(strict=False)
        for pdf in file_paths:
            file_merger.append(open(pdf, 'rb'))  # 合并pdf文件
        with open(selected_path + '\发票_合并.pdf', 'wb') as fout:
            file_merger.write(fout)

    return list,num


# def ticket_classify(unusedfilepath,excelsiorpath)
#     #读取表格中的发票形成列表
#     stexcel = pd.read_excel(excelsiorpath)
#
#     ex_value = stexcel.values
#     len_r = len(ex_value)
#     data = []
#     for i in range(len_r):
#         data.append(ex_value[i][3])
#     print(data)
def RollBack(Inovice_ID, UnuseFilepath):
    flag = 0
    hasname = []  # 列表
    findpath = UnuseFilepath + "/已报销"
    for item in Inovice_ID:  # 存入每一个待回滚的发票路径到hasname
        hasname.append(item + prefix1 + suffix)

    files = os.listdir(findpath)  # 得到文件夹下的所有文件名称
    for file in files:  # 遍历文件夹
        if file in hasname:
            flag = 1
            temp = findpath + os.sep + file
            shutil.move(temp, UnuseFilepath)

            list="回滚成功！"
    if flag == 0:
        list="未找到需要回滚的发票，请检查您输入的发票代码！"
    return list


def read(filepath):
    rexcel1 = seek_creat_files(filepath, "ExcelStorage")
    rexcel1 = rexcel1 + "/excelsior.xlsx"
    for kind in kinds:
        newfilepath = rpdf = filepath + "/" + kind
        check_and_creat_dir(newfilepath)
    return rexcel1


def select(str_in, filepath, rexcel, money):
    typenum = [int(n) - 1 for n in str_in.split()]
    # 输入未报销发票的文件夹地址，和存储未报销发票的excel文件地址，以及类别序号，进行数据库录入，以及去重排序
    UnuesdTicket_produceexcel_throw_sort(filepath, rexcel, typenum)
    throw_repeat(rexcel, "Sheet")  # 去重
    excel_sort(rexcel, "Sheet")  # 排序
    # 根据最大金额，读取未报销的excel文件，给出最优报销
    list,num=excel_MC_combination(filepath,rexcel, money)
    return list,num


# 主文件
if __name__ == "__main__":
    # 未报销的发票文件夹目录 用户输入
    print("输入未报销发票存储文件夹地址。")
    print("举例：D:\桌面\ text")
    UnuseFilepath = input("请输入：")
    print("选择你需要的工作模式：正常报销——1，发票回滚——2.")
    mode = input("请输入：")
    if mode == '1':
        print("您输入的未报销发票存储文件夹地址为: ", UnuseFilepath)
        print("我们即将在您输入的文件夹内建立两个文件夹 “已报销” 和 “未报销“ 分别存储对应的发票")
        print("我们将在您输入的文件夹下创建存储excel的文件夹，名为 “ExcelStorage” ")
        rexcel = seek_creat_files(UnuseFilepath, "ExcelStorage")
        rexcel = rexcel + "/excelsior.xlsx"

        # 创建文件夹（已报销与未报销）
        for kind in kinds:
            newfilepath = rpdf = UnuseFilepath + "/" + kind
            check_and_creat_dir(newfilepath)
        print(" ")
        print("******************日常报销的票据类别******************")
        print("1.交通费: 包括市内公交车费，出租车费、市内汽油费等。")
        print("2.差旅费: 包括餐饮发票和住宿发票等。")
        print("3.耗材费: 包括办公耗材发票，办公维修及配件发票。通常指的是办公室OA办公设备、IT和数码设备日常运作、维修、维护所需要的材料。")
        print("4.印刷费: 包含资料冲印费、资料复印费、图文编辑费、信封印刷费等。")
        print("5.培训费: 教职工或公司员工培训费用，包括国内培训费和国外培训费。")
        print("6.其他  : 上述几类中所不包含的发票。")
        print("请输入本次您需要报销的发票类别：（只需输入序号，如果是多类别序号用空格隔开，如：1 3）")
        str_in = input()
        typenum = [int(n) - 1 for n in str_in.split()]

        # 输入未报销发票的文件夹地址，和存储未报销发票的excel文件地址，以及类别序号，进行数据库录入，以及去重排序
        UnuesdTicket_produceexcel_throw_sort(UnuseFilepath, rexcel, typenum)
        throw_repeat(rexcel, "Sheet")  # 去重
        excel_sort(rexcel, "Sheet")  # 排序
        # 根据最大金额，读取未报销的excel文件，给出最优报销
        excel_MC_combination(rexcel)

    # Update_excel(UnuseFilepath + "/未报销", rexcel)
    elif mode == '2':
        print("输入需要回滚的发票代码（注意不是发票号码）,如需回滚多张请用英文逗号隔开.例如：1234566,1234567")
        Inovice_ID = input("请输入：").split(',')
        RollBack(Inovice_ID, UnuseFilepath)
    else:
        print("输入错误，END！")
